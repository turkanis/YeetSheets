#!/bin/env python3

import argparse
import csv
import json
import os
import re
from openpyxl import load_workbook
import sys

MATCH_CELL_COORDS = re.compile(r'^([A-Z]:[1-9][0-9]*)(?:\((.+)\))?$')

# Parses the command line for this tool
def parseCommandLine():

    # Parse command line
    cmd = argparse.ArgumentParser(
             prog='process-countries',
             description='Applies a template to each sheet in an XSLX document and outputs the result as TSV',
          )
    cmd.add_argument(
        '-t', '--template',
        dest='template',
        help='Path to XLSX template'
    )
    cmd.add_argument(
        '-c', '--countries',
        dest='countries',
        help='Path to XLSX document containing one sheet per county'
    )
    args = cmd.parse_args(sys.argv[1:])

    # Validate options
    if args.template == None:
        outputError("Missing template\n")
    checkFile(args.template)
    if args.countries == None:
        outputError("Missing country data\n")
    checkFile(args.countries)

    return args

def checkFile(file):
    if not os.path.exists(file):
        outputError("No such file: %s\n" % file)
    if not os.path.isfile(file):
        if os.path.isdir(file):
            outputError("The file '%s' is a directory\n" % file)
        else:
            outputError("The file '%s' is not a plain file\n" % file)

def outputError(msg):
    sys.stderr.write(msg)
    exit(1)

# Returns a dictionary mapping column labels to cell coordinates of the form
# X:N
def loadTemplate(file):
    wb = load_workbook(filename = file)
    data = []
    for row in wb.active.values:
        data.append([])
        for value in row:
            data[-1].append(value)
    if len(data) != 2:
        raise RuntimeError("Expected 2 rows in template: found %i" % len(data))
    if len(data[0]) != len(data[1]):
        raise RuntimeError("Number of column labels in template does not match number of cell coordinates")
    template = {}
    for i in range(len(data[0])):
        label = data[0][i]
        spec = data[1][i]
        if spec == None:
            break
        match = MATCH_CELL_COORDS.search(spec)
        if not match:
            raise RuntimeError("Malformed cell coordinates in template at column %s: %s" % (i, spec))
        (coords, format) = match.group(1, 2);
        template[label] = { 'coords' : coords, 'format' : format }
    return template

# Returns a dictionary mapping country names to dictionaries mapping cell
# coordinates of the for X:N to cell values
def loadCountries(file):
    print("Loading countries workbook", flush=True, file=sys.stderr)
    wb = load_workbook(filename = file)
    print("Countries workbook loaded", flush=True, file=sys.stderr)
    countries = {}
    for sheet in wb:
        print("Sheet = %s" % sheet.title, flush=True, file=sys.stderr)
        country = sheet.title
        countries[country] = {}
        rowNum = 1
        values = {}
        for row in sheet.values:
            colNum = 0
            for value in row:
                countries[country]['%s:%i' % (columnNames()[colNum], rowNum)] = value
                colNum += 1
            rowNum += 1
    return countries

# Applies the given template, of the form returned by loadTemplate(), to
# the country data returned by loadCountries(), producing a two-dimensional
# array with a header row followed by one row for each country
def applyTemplate(template, countries):
    results = []

    # Create header row
    header = list(template.keys())
    unique = set()
    for v in header:
        if v in unique:
            raise RuntimeError("Duplicate label %s in template" % v)
        unique.add(v)

    results.append(header)

    # Create one row for each country
    for country, values in countries.items():
        row = []
        for spec in template.values():
            coords = spec['coords']
            if not coords in values:
                raise RuntimeError("No data for cell %s in country %s" % (coords, country))
            value = values[coords]
            if spec['format']:
                try:
                    print("Applying format %s" % spec['format'], flush=True, file=sys.stderr)
                    value = spec['format'].format(value)
                except:
                    print("WARNING: Format error for %s at %s: %s" % (country, coords, sys.exc_info()[1]), flush=True, file=sys.stderr)
            row.append(value)
        results.append(row)
    return results

COLUMN_NAMES = None
def columnNames():
    global COLUMN_NAMES
    if COLUMN_NAMES == None:
        COLUMN_NAMES = []
        alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        for a in alpha:
            COLUMN_NAMES.append(a)
        for a in alpha:
            for b in alpha:
                COLUMN_NAMES.append(a + b)
    return COLUMN_NAMES

# Writes the results to a tab-delimited values file
def writeResults(results):
    writer = csv.writer(sys.stdout, delimiter="\t", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    for row in results:
        writer.writerow(row)

# Script body
args = parseCommandLine()
template = loadTemplate(args.template)
countries = loadCountries(args.countries)
results = applyTemplate(template, countries)
writeResults(results)

