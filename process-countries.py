#!/bin/env python3.6

import argparse
import csv
import json
from os import path
import re
from openpyxl import load_workbook
import sys

TEMPLATE = "template.xlsx"
RESULTS_FILE = "results.tsv"
DATA = "countries.xlsx"
MATCH_CELL_COORDS = re.compile(r'^[A-Z]:[1-9][0-9]*$')
IMAGE_COLUMN = 3
IMAGE_FORMAT="/Links/%s.svg"
NOTE_LABEL = "NOTE"
FOOTNOTE_LABEL = "FOOT"
LAST_DATA_ROW = 69
MAX_NOTES = 10
MAX_FOOTNOTES = 10

# Parses the command line for this tool
def parseCommandLine():

    # Parse command line
    cmd = argparse.ArgumentParser(
             prog='process-countries',
             description='Applies a template to each sheet in an XSLX document and outputs the result as TSV',
          )
    cmd.add_argument(
        '-t', '--template',
        dest='template',
        help='Path to XLSX template'
    )
    cmd.add_argument(
        '-c', '--countries',
        dest='counties',
        help='Path to XLSX document containing one sheet per county'
    )
    args = cmd.parse_args(sys.argv[1:])

    # Validate options
    if args.template == None:
        outputError("Missing template\n")
    checkFile(args.template)
    if args.countries == None:
        outputError("Missing country data\n")
    checkFile(args.countries)

    return args

def checkFile(file):
    if not os.path.exists(file):
        outputError("No such file: %s\n" % file)
    if not os.path.isfile(file):
        if os.path.isdir(file):
            outputError("The file '%s' is a directory\n" % file)
        else:
            outputError("The file '%s' is not a plain file\n" % file)

def outputError(msg):
    sys.stderr.write(msg)
    exit(1)

# Returns a dictionary mapping column labels to cell coordinates of the form
# X:N
def loadTemplate(file):
    wb = load_workbook(filename = file)
    data = []
    for row in wb.active.values:
        data.append([])
        for value in row:
            data[-1].append(value)
    if len(data) != 2:
        raise RuntimeError("Expected 2 rows in template: found %i" % len(data))
    if len(data[0]) != len(data[1]):
        raise RuntimeError("Number of column labels in template does not match number of cell coordinates")
    template = {}
    for i in range(len(data[0])):
        coords = data[1][i]
        if not MATCH_CELL_COORDS.search(coords):
            raise RuntimeError("Malformed cell coordinates in template at column %s: %s" % (i, coords))
        template[data[0][i]] = data[1][i];
    return template

# Returns a dictionary mapping country names to dictionaries mapping cell
# coordinates of the for X:N to cell values
def loadCountries(file):
    print("Loading countries workbook", flush=True)
    wb = load_workbook(filename = file)
    print("Countries workbook loaded", flush=True)
    countries = {}
    for sheet in wb:
        print("Sheet = %s" % sheet.title, flush=True)
        country = sheet.title
        countries[country] = {}
        rowNum = 1
        values = {}
        for row in sheet.values:
            colNum = 0
            for value in row:
                countries[country]['%s:%i' % (columnNames()[colNum], rowNum)] = value
                colNum += 1
            rowNum += 1
    return countries

# Applies the given template, of the form returned by loadTemplate(), to
# the country data returned by loadCountries(), producing a two-dimensional
# array with a header row followed by one row for each country
def applyTemplate(template, countries):
    results = []

    # Create header row
    header = list(template.keys())
    unique = set()
    for v in header:
        if v in unique:
            raise RuntimeError("Duplicate label %s in template" % v)
        unique.add(v)

    for i in range(1, MAX_NOTES + 1):
        header.append("NOTE%i" % i)
    for i in range(1, MAX_FOOTNOTES + 1):
        header.append("FOOTNOTE%i" % i)
    results.append(header)

    # Create one row for each country
    for country, values in countries.items():
        row = []
        for coords in template.values():
            if not coords in values:
                raise RuntimeError("No data for cell %s in country %s" % (coords, country))
            row.append(values[coords])
        processNotes(country, values, row)
        formatImageLink(country, values, row)
        results.append(row)
    return results

# Adds entries for notes and footnotes to target row, using values from source
def processNotes(country, source, target):
    rowNum = LAST_DATA_ROW + 1

    # Skip blank rows before notes:
    while getColA(source, rowNum) == None:
        rowNum += 1

    # Add notes to target
    notes = []
    while len(notes) < MAX_NOTES:
        value = getColA(source, rowNum)
        if value == None:
            break
        else:
            notes.append(value)
            rowNum += 1
    if getColA(source, rowNum) != None:
        raise RuntimeError("More than %i notes for country %s" % (MAX_NOTES, country))
    while len(notes) < MAX_NOTES:
        notes.append(None)
    target += notes

    # Skip blank rows before footnotes:
    while getColA(source, rowNum) == None:
        rowNum += 1

    # Add footnotes to target
    footnotes = []
    while len(footnotes) < MAX_NOTES:
        value = getColA(source, rowNum)
        if value == None:
            break
        else:
            footnotes.append(value[0] + "\t" + value[2:])
            rowNum += 1
    if getColA(source, rowNum) != None:
        raise RuntimeError("More than %i footnotes for country %s" % (MAX_NOTES, country))
    while len(footnotes) < MAX_NOTES:
        footnotes.append(None)
    target += footnotes

COLUMN_NAMES = None
def columnNames():
    global COLUMN_NAMES
    if COLUMN_NAMES == None:
        COLUMN_NAMES = []
        alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        for a in alpha:
            COLUMN_NAMES.append(a)
        for a in alpha:
            for b in alpha:
                COLUMN_NAMES.append(a + b)
    return COLUMN_NAMES

# Helper for processNotes()
def getColA(values, rowNum):
    coords = "A:%i" % rowNum
    return values[coords] if coords in values else None

# Formats the image link in column D
def formatImageLink(country, source, target):
    target[IMAGE_COLUMN] = IMAGE_FORMAT % country

# Writes the results to a tab-delimited values file
def writeResults(results):
    #with open(RESULTS_FILE, mode='w') as file:
    writer = csv.writer(sys.stdout, delimiter="\t", quotechar='"', quoting=csv.QUOTE_MINIMAL)
    for row in results:
        writer.writerow(row)

# Script body
args = parseCommandLine()
template = loadTemplate(args.template)
countries = loadCountries(args.countries)
results = applyTemplate(template, countries)
writeResults(results)

